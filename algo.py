import copy
import pandas as pd
import sys 
import json
from openpyxl import load_workbook
from datetime import date
from datetime import datetime
import datetime as og
import holidays
hope = holidays.Ghana()

#to be used for classroom assignment tree
class Node:
    def __init__(self, options, value, steps):
        self.options = options
        self.value = value
        self.steps = steps
        self.children = []
    def get_options(self):
        return self.options
    def get_value(self):
        return self.value
    def get_steps(self):
        return self.steps
    def birth(self):
        for x in self.options:
            x= int(x)
            new_value = self.value - x
            new_steps = self.steps + [x]
            check = len(self.options[x])
            for_kids = copy.deepcopy(self.options)

            if (check <= 1):
                del for_kids[x]
            else:
                for_kids[x].pop()
            
            self.children.append(Node(for_kids, new_value, new_steps)) 

        return self.children

#to be used for tree building for classroom assignment
class tree_builder:
    def __init__(self, classrooms, value):
        self.head = Node(classrooms, value, [])
        self.level = [self.head]

    def build(self):
        z = []
        values = {}
        val = False
        for x in self.level:
            q = x.birth()
            for y in q:
                if y.get_value() not in values:
                    z.append(y)
                    values[y.get_value()] = "here"
                    if(y.get_value() <= 0):
                        val = True
        self.level= z
        return [self.level, val]
#main class
class Course:
    def __init__(self, name):
        #this variable represents the name of the course
        self.name = name
        #this variable represents the students taking that course
        self.students = []
        self.red_flags = []
        #this represents the courses that this course cannot be grouped with
        '''this variable represents the students taking the course as a major course 
        elective and is to be used in implementing the soft constraints'''
        self.majors = 0
        self.size = None
        self.color = 9003
        self.classrooms = []
        #this is the breakdown of the different year groups of students taking this coursec, 
        # ordered from largest to smallest year group proportion 
        self.breakdown = []
        self.options = 0 

    #mutator methods for the course attributes
    def set_options(self, options):
        self.options = options
    def  set_students(self, student):
        self.students.append(student)
    def set_breakdown(self, breakdown):
        self.breakdown = breakdown
    def  set_majors(self):
        self.majors += 1
    def set_name(self, name):
        self.name = name
    def update_classrooms(self, name):
        self.classrooms.append(name)
    def set_size(self, size):
        self.size = size
    def update_red(self, red):
        #update red flags using a single course
        for x in red:
            if x not in self.red_flags and x != self.name:
                self.red_flags.append(x)
    def set_red(self, red):
        #update red flags using an entire list
        self.red_flags = red
    def set_color(self, color):
        #to be used for graph colouring
        self.color = color

    #accessor methods
    def get_breakdown(self):
        return self.breakdown;
    def get_breakdown(self):
        return self.breakdown;
    def get_color(self):
        return self.color
    def get_classrooms(self):
        return self.classrooms
    def get_red(self):
        return self.red_flags
    def get_name(self):
        return self.name
    def get_students(self):
        return self.students
    def get_majors(self):
        return self.majors
    def get_size(self):
        self.set_size(len(self.students))
        return len(self.students)
    def get_options(self):
        return self.options
    
    #function to check if a course is compatible with a list of other courses
    def is_compatible(self, others):
        if not others:
            return True
        for x in others:
            #not compatible
            if x.get_name() in self.red_flags: 
                return False
        return True
    def get_compatibility_rating(self, others, courses):
        '''
        Function to get the level of compatibility a course has with a group of other courses
        returns the number of overlapping students between the course and the group
        and the names of the courses with whom it has overlapping students
        '''
        overlap = list(set([x.get_name() for x in others]) & set(self.red_flags)) 
        num_clashing_students = 0
        for x in overlap:
            num_clashing_students += len(set(courses[course_index_hash_map[x]].get_students()) & set(self.get_students()))
        return [num_clashing_students , overlap]
        
   
class Student:
    def __init__(self, ID, course):
        #this variable represents the student's id
        #these represent is the student's program and year group
        self.id = ID
        self.course = course        
        self.year_group = ID[-4:]
        self.enrolled = []
    def set_enrolled(self, enrolled):
        self.enrolled = enrolled
    def get_ID(self):
        return self.id
    def get_course(self):
        return self.course
    def get_year_group(self):
        return self.year_group
        
def tree_user(classrooms, value): 
    space =   sum([x * len(classrooms[x]) for x in classrooms])
    if(space < value):
        print("VALUE TOO LARGE",space, value)
        return False 
    z = tree_builder(classrooms, value)
    x = []
    nonzero = True

    while nonzero:
        x = z.build()
        nonzero = not x[1] 
    min_val = -98765434567
    min_node = None
    for node in x[0]:
        if node.get_value() > min_val and node.get_value() <= 0:
            if min_node != None and len(node.get_steps()) >= len(min_node.get_steps()):
                pass                 
            else:       
                min_val = node.get_value()
                min_node = node
    return(min_node.get_steps())

class Slot:
    '''
    These objects represent unique time slots for example 8am, 1pm, etc
    '''
    def __init__(self, classrooms):
        #this represents the courses to be examined in that slot
        self.courses = []      
        #the classrooms available for use 
        self.classrooms = classrooms
        #the total space avaible 
        self.available_space = sum([x * len(self.classrooms[x]) for x in self.classrooms])
    #accessor methods
    def get_courses(self):
        return self.courses
    def get_available_space(self):
        return self.available_space
    #method to assign a new course to a slot.
    #assign is only called if there is enough available space
    def assign(self, course):
        size = course.get_size()
        course_classrooms = tree_user(self.classrooms, size)
        if (course_classrooms):
            self.courses.append(course)
            for x in course_classrooms:
                room = self.classrooms[x][0]
                #give course classroom update_classrooms()
                course.update_classrooms(room)
                #reduce available space
                self.available_space -= x
                # reduce self classrooms
                if(len(self.classrooms[x]) == 1 ):
                    del self.classrooms[x]
                else:
                    self.classrooms[x].pop(0)                 

def scheduler(exam_period, courses):
    ''' 
    This function takes a variable representing the number of days available for examinations and 
    the list of course objects to be scheduled and 
    Returns:
        set of scheduled and unscheduled courses
    '''
    unscheduled = []
    scheduled = []
    my_queue = []
    visited = []
    start = courses[0]
    my_queue.append(start)
    while my_queue:
        current = my_queue.pop(0)
        reds = [courses[course_index_hash_map[x]] for x in current.get_red()]
        colors = set([y.get_color() for y in reds if y in visited])
        for x in range(0, exam_period):
            if x not in colors:
                current.set_color(x)
                scheduled.append(current)
        visited.append(current)
        if current.get_color() not in range(0, exam_period):
            unscheduled.append(current)            
        my_queue.extend([y for y in courses if y not in reds and y not in visited and y not in my_queue])
        my_queue.sort(key=lambda x: len([y for y in x.get_red() if y not in visited]), reverse=True)
    return (scheduled, unscheduled)

def re_scheduler(exam_period, courses, to_schedule):
    ''' 
    Schedules only the courses in `to_schedule` list over the given `exam_period` days,
    while considering conflicts (reds) based on the full `courses` list.
    
    Args:
        exam_period (int): Number of days available for exams.
        courses (list): All course objects (needed for conflict resolution).
        to_schedule (list): Subset of course objects to be scheduled.
        
    Returns:
        tuple: (scheduled_courses, unscheduled_courses)
    '''
    unscheduled = []
    scheduled = []
    my_queue = []
    visited = []
    
    # Start with the first course to be scheduled
    start = to_schedule[0]
    my_queue.append(start)
    
    while my_queue:
        current = my_queue.pop(0)
        
        # Get conflicting courses (reds) using full courses list
        reds = [courses[course_index_hash_map[x]] for x in current.get_red()]
        colors = set([y.get_color() for y in reds if y in visited])
        
        for x in range(0, exam_period):
            if x not in colors:
                current.set_color(x)
                scheduled.append(current)
                break  # assign first available color and break
        
        visited.append(current)
        
        if current.get_color() not in range(0, exam_period):
            unscheduled.append(current)            

        # Add only the remaining to_schedule courses not yet visited
        my_queue.extend([y for y in to_schedule if y not in reds and y not in visited and y not in my_queue])
        my_queue.sort(key=lambda x: len([y for y in x.get_red() if y not in visited]), reverse=True)

    return (scheduled, unscheduled)


def find_largest_compatible_subset(courses, course_list):
    """
    Finds the largest set of non-conflicting courses.
    Prioritizes courses with fewer scheduling options.
    
    Args:
        courses (list): List of course objects.
        
    Returns:
        list: Compatible subset of courses.
    """
    conflict_map = {course: course.get_red()
                    for course in courses}

    # Sort courses by number of options (lowest first)
    sorted_courses = sorted(courses, key=lambda c: c.get_options())
    compatible_set = []
    used_courses = set()
    
    for course in sorted_courses:
        if course in used_courses:
            continue
        # Check if this course conflicts with any already selected
        if all(conflict not in compatible_set for conflict in conflict_map[course]):
            compatible_set.append(course)
            used_courses.update(conflict_map[course])
            used_courses.add(course)  # Mark itself as used too
    return compatible_set

def excel_to_csv(excel_file_name):
    file_as_df = pd.read_excel(excel_file_name, header=0)
    return file_as_df

def drop_unneccessary_columns(enrol_df):
    return enrol_df[enrol_df['Type'] == "Schedule"] 

def prep_student_and_courses(enrol_excel_name):
    #opening the file as a panda dataframes and 
    #dropping all unnecessary courses that are not being scheduled
    enrol_df = drop_unneccessary_columns(excel_to_csv(enrol_excel_name))
    enrol_df["year_program"] = enrol_df["Student Program"] + " "+ enrol_df["Student ID"].astype(str).str[-4:]
    enrol_df["year_program"] = enrol_df["Student Program"] + " "+ enrol_df["Student ID"].astype(str).str[-4:]
    #initializing the list of Course objects to be scheduled
    courses = []
    students = []
    index = 0 
    #setting up course objects
    courses_in_set = enrol_df["Course Name"].unique()
    for x in courses_in_set:
        #initialising the course
        new_course = Course(x)
        course_index_hash_map[x] = index
        index+=1
        breakdown = enrol_df[enrol_df["Course Name"] == x]["year_program"].value_counts().index.to_list()
        new_course.set_breakdown(breakdown)
        courses.append(new_course)

    #Setting up student objects
    students = enrol_df["Student ID"].unique()
    students = enrol_df["Student ID"].unique()    
    for id in students:
        #get all the rows associated with the student
        student_set = enrol_df[enrol_df["Student ID"] == id]
        #identify the student's program
        student_program = student_set["Student Program"].iloc[[0]]
        #use that data to create the student object
        student = Student(id, student_program)
        #get all the courses the student is enrolled in 
        enrolled =  student_set["Course Name"].unique().tolist()
        major_courses = student_set[student_set["Course Sub Category"] == "Required Major Classes"]["Course Name"].unique().tolist()
        #update course records as necessary.   
        for x in enrolled:
            course_loc = course_index_hash_map[x]
            #add student to the course's student list
            courses[course_loc].set_students(student)
            if x in major_courses:
                #update the courses's major var, if the student is taking it as a required major course
                courses[course_loc].set_majors()
            #update the red flags list in all courses to prevent clashing
            courses[course_loc].update_red(enrolled)

    courses.sort(key=lambda x: len(x.get_red()), reverse=True)
    index = 0
    for x in courses:
        course_index_hash_map[x.get_name()] = index
        index += 1

    return (courses, students, course_index_hash_map)

def prep_classroom_data(class_excel_name):
    classrooms = {}
    classes = excel_to_csv(class_excel_name)
    room_sizes = classes['Capacity'].unique().tolist()
    for x in room_sizes:
        classrooms[x] = classes[classes["Capacity"] == x]["Name"].unique().tolist()
        classrooms[x] = classes[classes["Capacity"] == x]["Name"].unique().tolist()
    return classrooms

def prep_output(courses, num_days):
    # printing result of scheduling
    output = [[] for _ in range (num_days)]
    for co in courses:
        if co.get_color() in range(num_days):
            output[co.get_color()].append(co) 
    return output
def go_over(list_output, unscheduled):
    for x in unscheduled:
        for y in range(len(list_output)):
            if (x.is_compatible(list_output[y])):
                list_output[y].append(x)
                unscheduled.remove(x)
    return list_output, unscheduled

def get_best_slot(unscheduled_courses, slots, courses, thresh):
        unscheduled_best = {}
        for x in unscheduled_courses:
            for y in range(len(slots)):
                num_clashes = x.get_compatibility_rating(slots[y], courses)
                if num_clashes[0] <= thresh:
                    if x.get_name() not in unscheduled_best:
                        unscheduled_best[x.get_name()] = []
                    unscheduled_best[x.get_name()].append([y, num_clashes[1], num_clashes[0]])
            x.set_options(len(unscheduled_best[x.get_name()]))
        return unscheduled_best

def order_best_slot(unscheduled_best, num_days):
    actual_best = [[[], []] for x in range(num_days)]
    for x in unscheduled_best:
        y = unscheduled_best[x]
        for z in y:
            #course that fit
            actual_best[int(z[0])][0].extend([x])
            #clashing
            if z[1][0] not in actual_best[z[0]][1]:
                actual_best[z[0]][1].extend(z[1])
    return actual_best
def classroom_assigner(classrooms, scheduled_courses, best_slots):
    final_day_slots = [[] for _ in range(len(scheduled_courses))]
    for x in range(len(scheduled_courses)):
        to_be_scheduled = scheduled_courses[x]
        clashing = best_slots[x][1]
        final_day_slots[x] = assignments(to_be_scheduled, classrooms, clashing)
    return final_day_slots

def assignments(cour, classrooms, clashing):
    eight_am = Slot(copy.deepcopy(classrooms))
    one_pm = Slot(copy.deepcopy(classrooms))
    #order courses by putting clashing courses first
    cour = sorted(cour, key=lambda x: x.get_name() not in clashing)  
    for x in cour:
        if (one_pm.get_available_space() >= x.get_size()):
            one_pm.assign(x)
        elif(eight_am.get_available_space() >= x.get_size()):
            eight_am.assign(x)
    return [eight_am, one_pm]
    

def individual_dfs(day, time, slot):
    daytimes = [day+ " "+ time for x in range(len(slot.get_courses()))]
    cours = [x.get_name() for x in slot.get_courses()]
    rooms = [x.get_classrooms() for x in slot.get_courses()]
    breakdowns = [x.get_breakdown() for x in slot.get_courses()]
    tests = pd.DataFrame( list(zip(daytimes, cours, rooms, breakdowns)), columns=["Day - Time", 'Courses', 'Location', 'Class/Major of Registered Students'])
    return tests
def order_excel(final, dates):
    og_df = pd.DataFrame()
    for x in range(len(final)):
        day = final[x]
        eight = individual_dfs(f"{dates[x]}","Eight AM",day[0])
        one = individual_dfs(f"{dates[x]}" , "One PM", day[1])

        og_df = pd.concat([og_df, eight,one], ignore_index=True)
    og_df['Class/Major of Registered Students'] = og_df['Class/Major of Registered Students'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    og_df['Location'] = og_df['Location'].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
 
    excel_path = "/var/www/html/cap/venv/bin/python /var/www/html/cap/final/final.xlsx"
    aight = og_df.to_excel(excel_path, index= False)
    wb = load_workbook(excel_path)
    ws = wb.active

    current_value = None
    start_row = 2  # Excel rows (1-indexed, with header at row 1)
    for i in range(2, len(og_df) + 2):  # 2 to 5
        cell_value = ws[f"A{i}"].value
        if cell_value != current_value:
            if current_value is not None and start_row != i - 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
            current_value = cell_value
            start_row = i
    # Final merge if last rows are the same
    if start_row != len(og_df) + 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=len(og_df) + 1, end_column=1)
    wb.save(excel_path)
def get_dates(start_date, num_days):
    date_c = start_date
    y = []
    x = 0
    while x < num_days:
        #if not a weekend or holdiay, add to list
        if(date_c.weekday() <= 4 and not(date_c in hope)):
            y.append(date_c.strftime("%d %B, %Y"))
            x+=1
            date_c = date_c + og.timedelta(days=1)
        else:
            date_c = date_c + og.timedelta(days=1)
    return y

def schedule_unscheduled(final_assignment, actual_best, courses):
    sch = []
    for num in range(len(actual_best)):
        x = actual_best[num]
        z = find_largest_compatible_subset([courses[course_index_hash_map[y]] for y in x[0] if y not in sch], courses)
        slots = final_assignment[num]
        for course in z:
            if slots[0].get_available_space() >= course.get_size() and len(set(course.get_red()) & set([values.get_name() for values in slots[0].get_courses()])) == 0:
                slots[0].assign(course)
                sch.append(course.get_name())
            elif slots[1].get_available_space() >= course.get_size() and len(set(course.get_red()) & set([vals.get_name() for vals in slots[1].get_courses()])) == 0:
                slots[1].assign(course)
                sch.append(course.get_name())
            else:
                pass

    return final_assignment
def order_groups(course_objects, num_days):
    my_list = [[] for x in range(num_days)]
    for x in course_objects[0]:
        my_list[x.get_color()].append(x)
    return my_list
def main(enrol_excel_name, classroom_excel_name, num_days, list_start_date):
    num_days = int(num_days)
    try:
        variables = prep_student_and_courses(enrol_excel_name)
    except ():
        print(json.dumps({
        "success": False,
        "message": "Invalid Student Enrollment Data! Check Hint",
        "export_path": ""
        }))

    courses = variables[0]
    students = variables[1]
    course_index_hash_map = variables[2]
    threshold_unscheduled = (int(len(courses)/num_days) * 3)/2 + 1

    try:
        classrooms = prep_classroom_data(classroom_excel_name)
    except:
        print(json.dumps({
        "success": False,
        "message": "Invalid Classroom Data! Check Hint",
        "export_path": ""
        }))

    hello = scheduler(num_days, courses)
    list_outs = prep_output(courses, num_days)
    final_out = go_over(list_outs, hello[1])
    best_slots = get_best_slot(final_out[1], final_out[0], courses, threshold_unscheduled)
    actual_best = order_best_slot(best_slots, num_days)
    final_assignment = classroom_assigner(classrooms, final_out[0], actual_best)
    final_fr = schedule_unscheduled(final_assignment, actual_best, courses)  
    dates = get_dates(datetime.strptime(list_start_date, "%Y-%m-%d").date(), num_days)
    final = order_excel(final_fr, dates)
    print(json.dumps({
        "success": True,
        "message": "All done!",
        "export_path": "final/final.xlsx"
        }))
    
course_index_hash_map = {}
if __name__ == "__main__":
    # # Make sure arguments are passed when running the script
    # if len(sys.argv) < 5:
    #     print("Usage: python algo.py <student_file> <classroom_file> <num_days> <formatted_date>")
    #     sys.exit(1)
    
    student_file = sys.argv[1]
    classroom_file = sys.argv[2]
    num_days = sys.argv[3]
    formatted_date = sys.argv[4]
    
    # Call main function
    main(student_file, classroom_file, num_days, formatted_date)
# main("files/original.xlsx", "files/classrooms.xlsx", 8, "2025-04-05")